# -*- coding: utf-8 -*-

import base64
import io
import logging
import re
from collections import defaultdict
from datetime import timedelta, timezone

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools.misc import formatLang

_logger = logging.getLogger(__name__)

_BARCODE_PREFIX_RE = re.compile(r'^\[\d{6,20}\]\s*')
_GLOBAL_INVOICE_RFC = 'XAXX010101000'
_GLOBAL_INVOICE_PARTNER_NAME = 'PUBLICO EN GENERAL'


class PosSession(models.Model):
    _inherit = 'pos.session'

    pharma_z_report_sent = fields.Boolean(
        string='Reporte Z farmaceutico enviado',
        readonly=True,
        copy=False,
    )

    def _validate_session(self, balancing_account=False, amount_to_balance=0, bank_payment_method_diffs=None):
        return super()._validate_session(
            balancing_account=balancing_account,
            amount_to_balance=amount_to_balance,
            bank_payment_method_diffs=bank_payment_method_diffs,
        )

    def _send_pharma_z_report_email(self):
        self.ensure_one()
        recipients = self._get_pharma_z_report_recipients()
        if not recipients:
            _logger.warning(
                "Reporte Z farmaceutico no enviado para %s: sin destinatarios.",
                self.name,
            )
            return False

        report = self._get_pharma_z_report_values()
        body_html = self.env['ir.qweb']._render(
            'bi_pos_stock.pharma_z_report_email_body',
            {'session': self, 'report': report},
            minimal_qcontext=True,
        )
        attachment = self._create_pharma_z_report_xlsx_attachment(report)
        subject = _("Reporte diario de ventas Z - %(session)s", session=self.name)
        self.env['mail.mail'].sudo().create({
            'subject': subject,
            'body_html': body_html,
            'email_to': recipients,
            'email_from': self._get_pharma_z_report_email_from(),
            'reply_to': self._get_pharma_z_report_reply_to(),
            'attachment_ids': [(4, attachment.id)],
            'auto_delete': False,
        }).send()
        self.sudo().pharma_z_report_sent = True
        self.message_post(body=_("Reporte Z farmaceutico con Excel adjunto enviado a %s.") % recipients)
        return True

    def action_send_pharma_z_report_preview_email(self):
        recipient = (self.env.user.partner_id.email or '').strip()
        if not recipient:
            raise UserError(_(
                "Tu usuario no tiene correo configurado. Agrega un email al contacto "
                "del usuario o prueba con una cuenta que tenga correo."
            ))

        sent = 0
        for session in self:
            report = session._get_pharma_z_report_values()
            body_html = self.env['ir.qweb']._render(
                'bi_pos_stock.pharma_z_report_email_body',
                {'session': session, 'report': report},
                minimal_qcontext=True,
            )
            attachment = session._create_pharma_z_report_xlsx_attachment(report, preview=True)
            self.env['mail.mail'].sudo().create({
                'subject': _("[PREVIEW] Reporte diario de ventas Z - %(session)s", session=session.name),
                'body_html': body_html,
                'email_to': recipient,
                'email_from': session._get_pharma_z_report_email_from(),
                'reply_to': session._get_pharma_z_report_reply_to(),
                'attachment_ids': [(4, attachment.id)],
                'auto_delete': False,
            }).send()
            session.message_post(body=_(
                "Vista previa del Reporte Z farmaceutico con Excel adjunto enviada a %s."
            ) % recipient)
            sent += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Reporte Z enviado"),
                'message': _(
                    "Se envio la vista previa del Reporte Z con Excel adjunto de %d sesion(es) a %s."
                ) % (sent, recipient),
                'type': 'success',
                'sticky': False,
            },
        }

    def _get_pharma_z_report_email_from(self):
        MailServer = self.env['ir.mail_server'].sudo()
        filter_email_from = self._get_pharma_z_report_email_from_filter(
            MailServer._get_default_from_filter()
        )
        if filter_email_from:
            return filter_email_from
        default_from = MailServer._get_default_from_address()
        return default_from or self.env.user.email_formatted or self.company_id.email

    def _get_pharma_z_report_email_from_filter(self, from_filter):
        for part in (from_filter or '').split(','):
            allowed_sender = part.strip()
            if not allowed_sender:
                continue
            if '@' in allowed_sender:
                return allowed_sender
            return 'noreply@%s' % allowed_sender
        return False

    def _get_pharma_z_report_reply_to(self):
        return (
            self.company_id.partner_id.email_formatted
            or self.env.user.email_formatted
            or self.company_id.email
            or False
        )

    def action_download_pharma_z_report_xlsx(self):
        if len(self) != 1:
            raise UserError(_("Selecciona una sola sesion POS para descargar el Excel."))
        self.ensure_one()
        report = self._get_pharma_z_report_values()
        attachment = self._create_pharma_z_report_xlsx_attachment(report, preview=True)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def action_print_cierre_caja_report(self):
        self.ensure_one()
        return self.env.ref('bi_pos_stock.action_report_corte_z').report_action(self)

    def _get_pharma_z_report_recipients(self):
        self.ensure_one()
        emails = [
            self.user_id.partner_id.email,
            self.env.user.partner_id.email,
            self.company_id.email,
        ]
        seen = set()
        recipients = []
        for email in emails:
            email = (email or '').strip()
            if not email or email.lower() in seen:
                continue
            seen.add(email.lower())
            recipients.append(email)
        return ','.join(recipients)

    def _get_pharma_z_report_values(self, orders=None):
        self.ensure_one()
        report_sessions = (
            (orders.mapped('session_id') or self).sudo()
            if orders is not None
            else self._get_pharma_z_report_scope_sessions()
        )
        orders = (
            orders.sudo()
            if orders is not None
            else self._get_pharma_z_report_orders_for_sessions(report_sessions)
        )
        lines = orders.mapped('lines')
        lot_cache = self._get_pharma_z_report_lot_cache(lines)

        movements = []
        order_groups = []
        for order in orders:
            local_order_dt = self._format_pharma_z_report_datetime(order.date_order)
            order_time = self._format_pharma_z_report_time(order.date_order)
            group_movements = []
            for line in order.lines.sorted('id'):
                group_movements.extend(
                    self._get_pharma_z_report_line_movements(
                        order=order,
                        line=line,
                        timestamp=local_order_dt,
                        lot_cache=lot_cache,
                    )
                )
            movements.extend(group_movements)
            group_total = sum(movement['amount_value'] for movement in group_movements)
            order_groups.append({
                'order_name': order.name or order.pos_reference or '',
                'customer_name': self._get_pharma_z_report_customer_name(order),
                'invoice_type': self._get_pharma_z_report_invoice_type(order),
                'timestamp': local_order_dt,
                'time': order_time,
                'line_count': len(group_movements),
                'total_value': group_total,
                'total': self._format_pharma_z_report_amount(group_total),
                'order_total_value': order.amount_total,
                'order_total': self._format_pharma_z_report_amount(order.amount_total),
                'lines': group_movements,
            })

        discount_count, discount_amount = self._get_pharma_z_report_discounts(lines)
        period_start, period_stop = self._get_pharma_z_report_period(report_sessions)
        total_qty = sum(movement['qty_value'] for movement in movements)
        ledger_total = sum(movement['amount_value'] for movement in movements)
        order_total = sum(orders.mapped('amount_total'))
        payments = self._get_pharma_z_report_payments(orders)
        payments_total_value = sum(orders.mapped('payment_ids').mapped('amount'))
        invoices = self._get_pharma_z_report_invoices(orders)
        invoices_total_value = sum(invoice['amount_value'] for invoice in invoices)
        taxes = self._get_pharma_z_report_taxes(orders)
        cash_control = self._get_pharma_z_report_cash_control(report_sessions, orders)
        refunds = self._get_pharma_z_report_refunds(orders)
        refunds_total_value = sum(refund['amount_value'] for refund in refunds)
        return {
            'company_name': self.company_id.name,
            'company_address': self._get_pharma_z_report_company_address(),
            'company_logo': self.company_id.logo,
            'session_name': ', '.join(report_sessions.mapped('name')) or self.name,
            'config_name': self.config_id.name,
            'start_at': self._format_pharma_z_report_datetime(period_start),
            'stop_at': self._format_pharma_z_report_datetime(period_stop),
            'sessions': self._get_pharma_z_report_session_rows(report_sessions, orders),
            'orders': self._get_pharma_z_report_order_rows(orders),
            'order_groups': order_groups,
            'movements': movements,
            'payments': payments,
            'payments_total': self._format_pharma_z_report_amount(payments_total_value),
            'payments_total_value': payments_total_value,
            'invoices': invoices,
            'invoice_count': len(invoices),
            'invoices_total': self._format_pharma_z_report_amount(invoices_total_value),
            'invoices_total_value': invoices_total_value,
            'taxes': taxes,
            'cash_control': cash_control,
            'cash_lines': self._get_pharma_z_report_cash_lines(report_sessions),
            'refunds': refunds,
            'refunds_total': self._format_pharma_z_report_amount(refunds_total_value),
            'refunds_total_value': refunds_total_value,
            'notes': self._get_pharma_z_report_notes(report_sessions),
            'discount_count': discount_count,
            'discount_amount': self._format_pharma_z_report_amount(discount_amount),
            'total_qty_value': total_qty,
            'total_qty': self._format_pharma_z_report_quantity(total_qty),
            'transaction_count': len(orders),
            'total_value': order_total,
            'total': self._format_pharma_z_report_amount(order_total),
            'ledger_total_value': ledger_total,
            'ledger_total': self._format_pharma_z_report_amount(ledger_total),
            'ledger_total_delta_value': order_total - ledger_total,
            'ledger_total_delta': self._format_pharma_z_report_amount(order_total - ledger_total),
            'warning_count': sum(1 for movement in movements if movement['warning']),
        }

    def _get_pharma_z_report_company_address(self):
        """Dirección de la empresa sin el nombre al inicio (evita duplicado con company_name)."""
        partner = self.company_id.partner_id
        address = partner.contact_address or ''
        name = (self.company_id.name or '').strip()
        # contact_address pone el nombre del partner en la primera línea; lo quitamos
        # para no repetirlo (el template ya muestra company_name en negrita).
        if name and address.startswith(name):
            address = address[len(name):].lstrip('\n\r ')
        return address

    def _get_pharma_z_report_orders(self):
        self.ensure_one()
        return self._get_pharma_z_report_orders_for_sessions(self._get_pharma_z_report_scope_sessions())

    def _get_pharma_z_report_orders_for_sessions(self, sessions):
        orders = self.env['pos.order'].sudo().search([
            ('session_id', 'in', sessions.ids),
            ('state', 'not in', ['draft', 'cancel']),
        ])
        return orders.sorted(key=lambda order: (order.date_order or fields.Datetime.now(), order.name or ''))

    def _get_pharma_z_report_scope_sessions(self):
        self.ensure_one()
        date_from, date_to = self._get_pharma_z_report_local_day_bounds(
            self.start_at or self.stop_at or fields.Datetime.now()
        )
        sessions = self.env['pos.session'].sudo().search([
            ('config_id', '=', self.config_id.id),
            ('start_at', '>=', fields.Datetime.to_string(date_from)),
            ('start_at', '<', fields.Datetime.to_string(date_to)),
        ], order='start_at asc, id asc')
        return sessions or self

    def _get_pharma_z_report_local_day_bounds(self, value):
        local_dt = fields.Datetime.context_timestamp(self, value)
        if local_dt.tzinfo is None:
            local_dt = local_dt.replace(tzinfo=timezone.utc)
        local_start = local_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        local_stop = local_start + timedelta(days=1)
        return (
            local_start.astimezone(timezone.utc).replace(tzinfo=None),
            local_stop.astimezone(timezone.utc).replace(tzinfo=None),
        )

    def _get_pharma_z_report_period(self, sessions):
        starts = [session.start_at for session in sessions if session.start_at]
        stops = [session.stop_at or session.start_at for session in sessions if session.stop_at or session.start_at]
        return (
            min(starts) if starts else self.start_at,
            max(stops) if stops else self.stop_at,
        )

    def _get_pharma_z_report_line_movements(self, order, line, timestamp, lot_cache):
        product = line.product_id
        lot_names = self._get_pharma_z_report_line_lot_names(line)
        if not lot_names:
            lot_names = [False]

        qty = line.qty or 0.0
        amount = (
            line.price_subtotal_incl
            if 'price_subtotal_incl' in line._fields
            else line.price_subtotal
        )
        discount_pct = line.discount or 0.0
        # Bruto antes del descuento × porcentaje = importe descontado por línea
        gross_per_unit = line.price_unit or 0.0
        discount_amount = abs(qty) * gross_per_unit * discount_pct / 100.0

        split_qty = self._get_pharma_z_report_split_qty(product, qty, lot_names)
        split_amount = amount / len(lot_names) if len(lot_names) > 1 else amount
        split_discount = discount_amount / len(lot_names) if len(lot_names) > 1 else discount_amount

        movements = []
        for lot_name in lot_names:
            lot = lot_cache.get((product.id, lot_name)) if lot_name and product else False
            expiration = self._get_pharma_z_report_lot_expiration(lot)
            requires_lot = product.tracking in ('lot', 'serial') if product else False
            warning = bool(requires_lot and (not lot_name or not expiration))
            movements.append({
                'timestamp': timestamp,
                'session_name': order.session_id.name or '',
                'order_name': order.name or order.pos_reference or '',
                'customer_name': self._get_pharma_z_report_customer_name(order),
                'invoice_type': self._get_pharma_z_report_invoice_type(order),
                'product_name': self._get_pharma_z_report_product_name(line),
                'lot_name': lot_name or (_('Sin lote') if requires_lot else ''),
                'expiration_date': expiration or (_('Sin caducidad') if lot_name and requires_lot else ''),
                'qty_value': split_qty,
                'qty': self._format_pharma_z_report_quantity(split_qty),
                'uom': product.uom_id.name if product and product.uom_id else '',
                'amount_value': split_amount,
                'amount': self._format_pharma_z_report_amount(split_amount),
                'discount_pct': discount_pct,
                'discount_amount_value': split_discount,
                'discount_amount': self._format_pharma_z_report_amount(split_discount) if split_discount else '',
                'warning': warning,
            })
        return movements

    def _get_pharma_z_report_line_lot_names(self, line):
        if 'pack_lot_ids' not in line._fields:
            return []
        lot_names = []
        for pack_lot in line.pack_lot_ids:
            name = pack_lot.lot_name if 'lot_name' in pack_lot._fields else False
            if not name and 'lot_id' in pack_lot._fields:
                name = pack_lot.lot_id.name
            if name and name not in lot_names:
                lot_names.append(name)
        return lot_names

    def _get_pharma_z_report_lot_cache(self, lines):
        lot_names = set()
        product_ids = set()
        for line in lines:
            if line.product_id:
                names = self._get_pharma_z_report_line_lot_names(line)
                if names:
                    product_ids.add(line.product_id.id)
                    lot_names.update(names)
        if not lot_names or not product_ids:
            return {}

        StockLot = self.env['stock.lot'].sudo()
        domain = [
            ('name', 'in', list(lot_names)),
            ('product_id', 'in', list(product_ids)),
        ]
        if 'company_id' in StockLot._fields:
            domain.append(('company_id', 'in', [self.company_id.id, False]))
        lots = StockLot.search(domain)
        return {
            (lot.product_id.id, lot.name): lot
            for lot in lots
        }

    def _get_pharma_z_report_lot_expiration(self, lot):
        if not lot or 'expiration_date' not in lot._fields or not lot.expiration_date:
            return False
        return self._format_pharma_z_report_date(lot.expiration_date)

    def _get_pharma_z_report_product_name(self, line):
        product = line.product_id
        name = (
            line.full_product_name
            if 'full_product_name' in line._fields and line.full_product_name
            else (product.display_name if product else '')
        )
        barcode = product.barcode if product else False
        if barcode:
            name = name.replace(f'[{barcode}] ', '').replace(f'[{barcode}]', '')
        return _BARCODE_PREFIX_RE.sub('', name or '').strip()

    def _get_pharma_z_report_customer_name(self, order):
        return order.partner_id.display_name if order.partner_id else _('Sin cliente')

    def _get_pharma_z_report_session_rows(self, sessions, orders):
        rows = []
        for session in sessions.sorted(key=lambda record: (record.start_at or fields.Datetime.now(), record.name or '')):
            session_orders = orders.filtered(lambda order: order.session_id == session)
            rows.append({
                'session_name': session.name,
                'cashier_name': session.user_id.name or '',
                'start_at': self._format_pharma_z_report_datetime(session.start_at),
                'stop_at': self._format_pharma_z_report_datetime(session.stop_at),
                'transaction_count': len(session_orders),
                'total': self._format_pharma_z_report_amount(sum(session_orders.mapped('amount_total'))),
            })
        return rows

    def _get_pharma_z_report_order_rows(self, orders):
        return [
            {
                'session_name': order.session_id.name or '',
                'order_name': order.name or order.pos_reference or '',
                'customer_name': self._get_pharma_z_report_customer_name(order),
                'invoice_type': self._get_pharma_z_report_invoice_type(order),
                'timestamp': self._format_pharma_z_report_datetime(order.date_order),
                'total': self._format_pharma_z_report_amount(order.amount_total),
            }
            for order in orders
        ]

    def _get_pharma_z_report_order_invoices(self, order):
        if 'account_move' not in order._fields:
            return self.env['account.move']
        moves = order.account_move
        return moves.filtered(lambda move: move.state != 'cancel')

    def _get_pharma_z_report_invoice_type(self, order):
        moves = self._get_pharma_z_report_order_invoices(order)
        if not moves:
            if self._is_pharma_z_report_pending_global_invoice(order):
                return _('Factura global pendiente')
            return _('Sin factura')
        if any(self._is_pharma_z_report_global_invoice(move) for move in moves):
            return _('Factura global')
        return _('Factura individual')

    def _is_pharma_z_report_pending_global_invoice(self, order):
        return (
            'make_global_invs' in order.session_id._fields
            and order.session_id.make_global_invs
            and 'invoice_status' in order._fields
            and order.invoice_status == 'to_invoice'
        )

    def _is_pharma_z_report_global_invoice(self, move):
        if 'is_global_invoice' in move._fields and move.is_global_invoice:
            return True
        partner = move.partner_id
        partner_vat = (partner.vat or '').strip().upper()
        partner_name = (partner.name or '').strip().upper()
        return partner_vat == _GLOBAL_INVOICE_RFC and partner_name == _GLOBAL_INVOICE_PARTNER_NAME

    def _get_pharma_z_report_split_qty(self, product, qty, lot_names):
        if len(lot_names) <= 1:
            return qty
        if product and product.tracking == 'serial':
            return 1.0 if qty >= 0 else -1.0
        return qty / len(lot_names)

    def _get_pharma_z_report_discounts(self, lines):
        count = 0
        amount = 0.0
        for line in lines:
            discount = line.discount or 0.0
            if not discount:
                continue
            count += 1
            amount += abs(line.qty or 0.0) * (line.price_unit or 0.0) * discount / 100.0
        return count, amount

    def _get_pharma_z_report_payments(self, orders):
        totals = defaultdict(float)
        for payment in orders.mapped('payment_ids'):
            totals[payment.payment_method_id.name or _('Sin metodo')] += payment.amount
        return [
            {
                'name': name,
                'amount': self._format_pharma_z_report_amount(amount),
            }
            for name, amount in sorted(totals.items())
        ]

    def _get_pharma_z_report_invoices(self, orders):
        grouped = {}
        for order in orders:
            moves = self._get_pharma_z_report_order_invoices(order)
            if not moves:
                continue
            amount_share = (order.amount_total or 0.0) / len(moves)
            for move in moves:
                if move.id not in grouped:
                    grouped[move.id] = {
                        'invoice_name': move.name,
                        'invoice_type': (
                            _('Factura global')
                            if self._is_pharma_z_report_global_invoice(move)
                            else _('Factura individual')
                        ),
                        'customer_name': move.partner_id.display_name,
                        'order_names': [],
                        'amount_value': 0.0,
                    }
                grouped[move.id]['amount_value'] += amount_share
                grouped[move.id]['order_names'].append(order.name or order.pos_reference or '')
        invoices = list(grouped.values())
        for invoice in invoices:
            invoice['order_names'] = ', '.join(dict.fromkeys(invoice['order_names']))
            invoice['amount'] = self._format_pharma_z_report_amount(invoice['amount_value'])
        return sorted(invoices, key=lambda invoice: invoice['invoice_name'] or '')

    def _get_pharma_z_report_invoices_total(self, orders):
        total = sum(
            invoice['amount_value']
            for invoice in self._get_pharma_z_report_invoices(orders)
        )
        return self._format_pharma_z_report_amount(total)

    def _get_pharma_z_report_refunds(self, orders):
        """Órdenes de devolución (total negativo) listadas por separado.

        Reproduce la sección #refunds del reporte nativo de detalles de ventas.
        Las devoluciones también quedan netadas en los totales del turno; aquí
        se exponen aparte para no omitir información del formato nativo.
        """
        refunds = []
        for order in orders:
            if (order.amount_total or 0.0) >= 0.0:
                continue
            refund_qty = sum(line.qty or 0.0 for line in order.lines)
            refunds.append({
                'order_name': order.name or order.pos_reference or '',
                'customer_name': self._get_pharma_z_report_customer_name(order),
                'time': self._format_pharma_z_report_time(order.date_order),
                'qty': self._format_pharma_z_report_quantity(refund_qty),
                'amount': self._format_pharma_z_report_amount(order.amount_total),
                'amount_value': order.amount_total or 0.0,
            })
        return refunds

    def _get_pharma_z_report_notes(self, sessions):
        """Notas de apertura/cierre de caja registradas en la(s) sesión(es)."""
        notes = []
        for session in sessions:
            opening = session.opening_notes if 'opening_notes' in session._fields else ''
            closing = session.closing_notes if 'closing_notes' in session._fields else ''
            opening = (opening or '').strip()
            closing = (closing or '').strip()
            if not opening and not closing:
                continue
            notes.append({
                'session_name': session.name or '',
                'opening_note': opening,
                'closing_note': closing,
            })
        return notes

    def _get_pharma_z_report_taxes(self, orders):
        """Desglose de IVA por tasa: base gravable e impuesto.

        Se calcula con ``compute_all`` sobre los impuestos de cada línea
        (aplicando posición fiscal cuando existe) para reproducir el resumen
        de impuestos que muestra el reporte nativo de detalles de ventas.
        """
        tax_totals = {}
        for order in orders:
            partner = order.partner_id or False
            currency = (
                order.session_id.currency_id
                or (order.currency_id if 'currency_id' in order._fields else False)
                or self.currency_id
                or self.company_id.currency_id
            )
            for line in order.lines:
                qty = line.qty or 0.0
                price_unit = (line.price_unit or 0.0) * (1.0 - (line.discount or 0.0) / 100.0)
                taxes = self._get_pharma_z_report_line_taxes(line)
                if not taxes:
                    base = (
                        line.price_subtotal_incl
                        if 'price_subtotal_incl' in line._fields
                        else line.price_subtotal
                    ) or 0.0
                    entry = tax_totals.setdefault(
                        0,
                        {
                            'name': _('Sin impuesto / Exento'),
                            'base': 0.0,
                            'amount': 0.0,
                        },
                    )
                    entry['base'] += base
                    continue
                computed = taxes.sudo().compute_all(
                    price_unit, currency, qty, product=line.product_id, partner=partner,
                )
                for tax_line in computed['taxes']:
                    entry = tax_totals.setdefault(
                        tax_line['id'],
                        {'name': tax_line['name'], 'base': 0.0, 'amount': 0.0},
                    )
                    entry['base'] += currency.round(tax_line.get('base', 0.0))
                    entry['amount'] += tax_line.get('amount', 0.0)

        rows = []
        base_total = 0.0
        amount_total = 0.0
        for entry in sorted(tax_totals.values(), key=lambda item: item['name']):
            base_total += entry['base']
            amount_total += entry['amount']
            rows.append({
                'name': entry['name'],
                'base': self._format_pharma_z_report_amount(entry['base']),
                'amount': self._format_pharma_z_report_amount(entry['amount']),
                'base_value': entry['base'],
                'amount_value': entry['amount'],
            })
        return {
            'rows': rows,
            'base_total': self._format_pharma_z_report_amount(base_total),
            'amount_total': self._format_pharma_z_report_amount(amount_total),
            'base_total_value': base_total,
            'amount_total_value': amount_total,
        }

    def _get_pharma_z_report_line_taxes(self, line):
        if 'tax_ids_after_fiscal_position' in line._fields:
            return line.tax_ids_after_fiscal_position
        if 'tax_ids' in line._fields:
            return line.tax_ids
        return self.env['account.tax']

    def _get_pharma_z_report_cash_lines(self, sessions=None):
        sessions = sessions or self
        cash_lines = []
        for session in sessions.sorted(key=lambda record: (record.start_at or fields.Datetime.now(), record.name or '')):
            for line in session.statement_line_ids.sorted(key=lambda record: (record.date or fields.Date.today(), record.id)):
                cash_lines.append({
                    'session_name': session.name,
                    'name': line.payment_ref or line.name or '',
                    'amount': self._format_pharma_z_report_amount(line.amount),
                })
        return cash_lines

    def _get_pharma_z_report_cash_control(self, sessions, orders):
        """Conciliación esperado/contado/diferencia por método de pago.

        Para los métodos en efectivo se usan los saldos de la caja registradora
        de la(s) sesión(es) (apertura, esperado y contado real). Para métodos
        electrónicos se toma el cierre real cuando Odoo generó diferencia de
        cierre o pagos bancarios asociados a la sesión.
        """
        sessions = sessions or self
        opening_total = 0.0
        for session in sessions:
            opening_total += self._get_pharma_z_report_field_amount(
                session,
                'cash_register_balance_start',
            )

        payment_groups = {}
        for payment in orders.mapped('payment_ids'):
            method = payment.payment_method_id
            if not method:
                continue
            session = self._get_pharma_z_report_payment_session(payment)
            key = (session.id if session else False, method.id)
            group = payment_groups.setdefault(
                key,
                {
                    'session': session,
                    'method': method,
                    'amount': 0.0,
                },
            )
            group['amount'] += payment.amount or 0.0

        method_rows = {}
        for group in payment_groups.values():
            method = group['method']
            session = group['session']
            key = method.id
            row = method_rows.setdefault(
                key,
                {
                    'name': method.name or _('Sin metodo'),
                    'method': method,
                    'expected_value': 0.0,
                    'counted_value': 0.0,
                    'difference_value': 0.0,
                    'session_ids': set(),
                },
            )
            expected, counted, difference = self._get_pharma_z_report_payment_control_values(
                session,
                method,
                group['amount'],
            )
            row['expected_value'] += expected
            row['counted_value'] += counted
            row['difference_value'] += difference
            if session:
                row['session_ids'].add(session.id)

        rows = []
        total_difference = 0.0
        currency = self.currency_id or self.company_id.currency_id
        for row in sorted(method_rows.values(), key=lambda value: value['name']):
            expected = row['expected_value']
            counted = row['counted_value']
            difference = row['difference_value']
            total_difference += difference
            rows.append({
                'name': row['name'],
                'expected': self._format_pharma_z_report_amount(expected),
                'counted': self._format_pharma_z_report_amount(counted),
                'difference': self._format_pharma_z_report_amount(difference),
                'difference_value': difference,
                'difference_class': (
                    'cz-num diff-ok'
                    if currency.is_zero(difference)
                    else 'cz-num diff-alert'
                ),
            })

        return {
            'opening': self._format_pharma_z_report_amount(opening_total),
            'opening_value': opening_total,
            'rows': rows,
            'total_difference': self._format_pharma_z_report_amount(total_difference),
            'total_difference_value': total_difference,
        }

    def _get_pharma_z_report_payment_session(self, payment):
        if 'session_id' in payment._fields and payment.session_id:
            return payment.session_id
        if 'pos_order_id' in payment._fields and payment.pos_order_id:
            return payment.pos_order_id.session_id
        return self.env['pos.session']

    def _get_pharma_z_report_payment_control_values(self, session, method, amount):
        if not session:
            return amount, amount, 0.0
        if self._is_pharma_z_report_cash_method(method):
            return self._get_pharma_z_report_cash_payment_control_values(
                session,
                amount,
            )
        difference = self._get_pharma_z_report_bank_payment_difference(
            session,
            method,
            amount,
        )
        return amount, amount + difference, difference

    def _get_pharma_z_report_cash_payment_control_values(self, session, amount):
        expected = self._get_pharma_z_report_field_amount(
            session,
            'cash_register_balance_end',
            default=None,
        )
        if expected is None:
            expected = (
                self._get_pharma_z_report_field_amount(session, 'cash_register_balance_start')
                + amount
                + self._get_pharma_z_report_field_amount(session, 'cash_real_transaction')
            )
        counted = self._get_pharma_z_report_field_amount(
            session,
            'cash_register_balance_end_real',
            default=expected,
        )
        return expected, counted, counted - expected

    def _get_pharma_z_report_bank_payment_difference(self, session, method, expected):
        if not method:
            return 0.0

        journal = method.journal_id if 'journal_id' in method._fields else False
        if journal:
            ref_value = "Closing difference in %s (%s)" % (method.name, session.name)
            move = self.env['account.move'].sudo().search([
                ('ref', '=', ref_value),
                ('journal_id', '=', journal.id),
            ], limit=1)
            if move:
                loss_account = (
                    journal.loss_account_id
                    if 'loss_account_id' in journal._fields
                    else False
                )
                is_loss = bool(
                    loss_account
                    and any(line.account_id == loss_account for line in move.line_ids)
                )
                return -move.amount_total if is_loss else move.amount_total

        AccountPayment = self.env['account.payment'].sudo()
        if not {'pos_session_id', 'pos_payment_method_id'} <= set(AccountPayment._fields):
            return 0.0
        account_payments = AccountPayment.search([
            ('pos_session_id', '=', session.id),
            ('pos_payment_method_id', '=', method.id),
        ])
        if not account_payments:
            return 0.0
        amount_field = 'amount_signed' if 'amount_signed' in AccountPayment._fields else 'amount'
        counted = sum(account_payments.mapped(amount_field))
        return counted - expected

    def _is_pharma_z_report_cash_method(self, method):
        return bool(
            method
            and 'is_cash_count' in method._fields
            and method.is_cash_count
        )

    @staticmethod
    def _get_pharma_z_report_field_amount(record, field_name, default=0.0):
        if field_name not in record._fields:
            return default
        return record[field_name] or 0.0

    def _create_pharma_z_report_xlsx_attachment(self, report, preview=False):
        self.ensure_one()
        filename = self._get_pharma_z_report_xlsx_filename(preview=preview)
        return self.env['ir.attachment'].sudo().create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(self._generate_pharma_z_report_xlsx(report)).decode('ascii'),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

    def _get_pharma_z_report_xlsx_filename(self, preview=False):
        self.ensure_one()
        session_name = re.sub(r'[^A-Za-z0-9_.-]+', '_', self.name or 'sesion').strip('_')
        prefix = 'PREVIEW_' if preview else ''
        return f'{prefix}Reporte_Z_{session_name}.xlsx'

    def _generate_pharma_z_report_xlsx(self, report):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise UserError(_(
                "No se puede generar el Excel del Reporte Z porque falta la libreria "
                "Python 'openpyxl' en el servidor."
            )) from exc

        workbook = Workbook()
        summary = workbook.active
        summary.title = 'Resumen'
        movements_sheet = workbook.create_sheet('Movimientos')

        header_fill = PatternFill('solid', fgColor='1F4E78')
        subheader_fill = PatternFill('solid', fgColor='D9EAF7')
        warning_fill = PatternFill('solid', fgColor='FFF2CC')
        white_font = Font(color='FFFFFF', bold=True)
        bold_font = Font(bold=True)
        money_format = '$#,##0.00'
        qty_format = '#,##0.####'

        summary['A1'] = 'Reporte diario de ventas Z'
        summary['A1'].font = Font(bold=True, size=14)
        summary.append([])
        for label, value in (
            ('Empresa', report['company_name']),
            ('Direccion', report['company_address']),
            ('Sesiones', report['session_name']),
            ('Punto de venta', report['config_name']),
            ('Periodo', f"{report['start_at']} - {report['stop_at']}"),
            ('Total', report['total']),
            ('Numero de transacciones', report['transaction_count']),
            ('Numero de descuentos', report['discount_count']),
            ('Importe de descuentos', report['discount_amount']),
            ('Advertencias de trazabilidad', report['warning_count']),
        ):
            summary.append([label, value])
            summary.cell(row=summary.max_row, column=1).font = bold_font

        self._append_pharma_z_report_xlsx_section(
            summary,
            'Sesiones / cajas',
            ['Sesion/Caja', 'Cajero', 'Fecha apertura', 'Fecha cierre', 'Transacciones', 'Total'],
            [
                [
                    session['session_name'],
                    session['cashier_name'],
                    session['start_at'],
                    session['stop_at'],
                    session['transaction_count'],
                    session['total'],
                ]
                for session in report['sessions']
            ],
            header_fill,
            white_font,
            Font(bold=True, size=12),
        )
        self._append_pharma_z_report_xlsx_section(
            summary,
            'Ordenes POS',
            ['Sesion/Caja', 'Orden POS', 'Cliente asociado', 'Tipo de factura', 'Fecha/Hora', 'Total'],
            [
                [
                    order['session_name'],
                    order['order_name'],
                    order['customer_name'],
                    order['invoice_type'],
                    order['timestamp'],
                    order['total'],
                ]
                for order in report['orders']
            ],
            header_fill,
            white_font,
            Font(bold=True, size=12),
        )
        self._append_pharma_z_report_xlsx_section(
            summary,
            'Pagos',
            ['Metodo', 'Importe'],
            [[payment['name'], payment['amount']] for payment in report['payments']],
            header_fill,
            white_font,
            Font(bold=True, size=12),
        )
        self._append_pharma_z_report_xlsx_section(
            summary,
            'Facturas',
            ['Factura', 'Tipo de factura', 'Cliente fiscal', 'Ordenes POS', 'Importe'],
            [
                [
                    invoice['invoice_name'],
                    invoice['invoice_type'],
                    invoice['customer_name'],
                    invoice['order_names'],
                    invoice['amount'],
                ]
                for invoice in report['invoices']
            ],
            header_fill,
            white_font,
            Font(bold=True, size=12),
        )
        self._append_pharma_z_report_xlsx_section(
            summary,
            'Movimientos de caja',
            ['Sesion/Caja', 'Referencia', 'Importe'],
            [[cash_line['session_name'], cash_line['name'], cash_line['amount']] for cash_line in report['cash_lines']],
            header_fill,
            white_font,
            Font(bold=True, size=12),
        )
        self._autosize_pharma_z_report_xlsx(summary)

        movement_headers = [
            'Fecha/Hora',
            'Sesion/Caja',
            'Orden POS',
            'Cliente asociado',
            'Tipo de factura',
            'Producto',
            'Lote',
            'Caducidad',
            'Cantidad',
            'Unidad',
            'Importe',
            'Descuento %',
            'Dto. Importe',
            'Advertencia',
        ]
        movements_sheet.append(movement_headers)
        for cell in movements_sheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')

        for movement in report['movements']:
            movements_sheet.append([
                movement['timestamp'],
                movement['session_name'],
                movement['order_name'],
                movement['customer_name'],
                movement['invoice_type'],
                movement['product_name'],
                movement['lot_name'],
                movement['expiration_date'],
                movement['qty_value'],
                movement['uom'],
                movement['amount_value'],
                movement['discount_pct'] or '',
                movement['discount_amount_value'] if movement.get('discount_pct') else '',
                _('Revisar lote/caducidad') if movement['warning'] else '',
            ])
            row = movements_sheet.max_row
            movements_sheet.cell(row=row, column=9).number_format = qty_format
            movements_sheet.cell(row=row, column=11).number_format = money_format
            if movement.get('discount_amount_value'):
                movements_sheet.cell(row=row, column=13).number_format = money_format
            if movement['warning']:
                for cell in movements_sheet[row]:
                    cell.fill = warning_fill

        total_row = movements_sheet.max_row + 1
        movements_sheet.cell(row=total_row, column=1, value='Total')
        movements_sheet.cell(row=total_row, column=1).font = bold_font
        movements_sheet.cell(row=total_row, column=9, value=sum(m['qty_value'] for m in report['movements']))
        movements_sheet.cell(row=total_row, column=9).font = bold_font
        movements_sheet.cell(row=total_row, column=9).number_format = qty_format
        movements_sheet.cell(row=total_row, column=11, value=sum(m['amount_value'] for m in report['movements']))
        movements_sheet.cell(row=total_row, column=11).font = bold_font
        movements_sheet.cell(row=total_row, column=11).number_format = money_format
        total_discount = sum(m.get('discount_amount_value', 0) for m in report['movements'])
        if total_discount:
            movements_sheet.cell(row=total_row, column=13, value=total_discount)
            movements_sheet.cell(row=total_row, column=13).font = bold_font
            movements_sheet.cell(row=total_row, column=13).number_format = money_format
        for cell in movements_sheet[total_row]:
            cell.fill = subheader_fill

        movements_sheet.freeze_panes = 'A2'
        movements_sheet.auto_filter.ref = movements_sheet.dimensions
        self._autosize_pharma_z_report_xlsx(movements_sheet)

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _append_pharma_z_report_xlsx_section(self, sheet, title, headers, rows, header_fill, white_font, title_font):
        if not rows:
            return
        sheet.append([])
        sheet.append([title])
        sheet.cell(row=sheet.max_row, column=1).font = title_font
        sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.fill = header_fill
            cell.font = white_font
        for row in rows:
            sheet.append(row)

    def _autosize_pharma_z_report_xlsx(self, sheet):
        for column_cells in sheet.columns:
            max_length = 0
            for cell in column_cells:
                value = '' if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            width = min(max(max_length + 2, 10), 48)
            sheet.column_dimensions[self._get_pharma_z_report_xlsx_column_letter(column_cells[0].column)].width = width

    def _get_pharma_z_report_xlsx_column_letter(self, number):
        letters = ''
        while number:
            number, remainder = divmod(number - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def _format_pharma_z_report_datetime(self, value):
        if not value:
            return ''
        tz = self.user_id.tz or self.env.user.tz
        local_dt = fields.Datetime.context_timestamp(self.with_context(tz=tz), value)
        return local_dt.strftime('%d/%m/%Y %H:%M:%S')

    def _format_pharma_z_report_time(self, value):
        if not value:
            return ''
        tz = self.user_id.tz or self.env.user.tz
        local_dt = fields.Datetime.context_timestamp(self.with_context(tz=tz), value)
        return local_dt.strftime('%H:%M')

    def _format_pharma_z_report_date(self, value):
        if not value:
            return ''
        if hasattr(value, 'date'):
            value = value.date()
        return value.strftime('%d/%m/%Y')

    def _format_pharma_z_report_quantity(self, value):
        return f'{value:g}'

    def _format_pharma_z_report_amount(self, value):
        return formatLang(self.env, value or 0.0, currency_obj=self.currency_id)


class ReportSaleDetails(models.AbstractModel):
    _inherit = 'report.point_of_sale.report_saledetails'

    @api.model
    def get_sale_details(self, date_start=False, date_stop=False, config_ids=False, session_ids=False, **kwargs):
        result = super().get_sale_details(
            date_start,
            date_stop,
            config_ids,
            session_ids,
            **kwargs,
        )
        movements = self._get_pharma_z_report_movements(
            date_start=date_start,
            date_stop=date_stop,
            config_ids=config_ids,
            session_ids=session_ids,
        )
        result.update({
            'pharma_z_report_movements': movements,
            'pharma_z_report_total_qty': sum(m['qty_value'] for m in movements),
            'pharma_z_report_total_amount': sum(m['amount_value'] for m in movements),
            'pharma_z_report_total_discount': sum(m.get('discount_amount_value', 0) for m in movements),
            'pharma_z_report_warning_count': sum(1 for m in movements if m['warning']),
        })
        return result

    def _get_pharma_z_report_movements(self, date_start=False, date_stop=False, config_ids=False, session_ids=False):
        sessions = self.env['pos.session'].browse(session_ids or []).exists()
        if sessions:
            movements = []
            for session in sessions:
                session_orders = session._get_pharma_z_report_orders_for_sessions(session)
                movements.extend(session._get_pharma_z_report_values(orders=session_orders)['movements'])
            return movements

        orders = self._get_pharma_z_report_orders_for_details(
            date_start=date_start,
            date_stop=date_stop,
            config_ids=config_ids,
        )
        movements = []
        for session in orders.mapped('session_id'):
            session_orders = orders.filtered(lambda order: order.session_id == session)
            movements.extend(session._get_pharma_z_report_values(orders=session_orders)['movements'])
        return movements

    def _get_pharma_z_report_orders_for_details(self, date_start=False, date_stop=False, config_ids=False):
        if hasattr(config_ids, 'ids'):
            config_ids = config_ids.ids
        date_start, date_stop = self._get_date_start_and_date_stop(date_start, date_stop)
        domain = [('state', 'in', ['paid', 'done'])]
        domain.append(('date_order', '>=', fields.Datetime.to_string(date_start)))
        domain.append(('date_order', '<=', fields.Datetime.to_string(date_stop)))
        if config_ids:
            domain.append(('config_id', 'in', config_ids))
        return self.env['pos.order'].sudo().search(
            domain,
            order='date_order asc, name asc, id asc',
        )


class ReportCorteZDocument(models.AbstractModel):
    _name = 'report.bi_pos_stock.report_corte_z_document'
    _description = 'Corte Z — Ventas por movimiento (PDF horizontal)'

    @api.model
    def _get_report_values(self, docids, data=None):
        sessions = self.env['pos.session'].browse(docids).exists()
        report_by_session = {
            session.id: session._get_pharma_z_report_values()
            for session in sessions
        }
        return {
            'doc_ids': sessions.ids,
            'doc_model': 'pos.session',
            'docs': sessions,
            'report_by_session': report_by_session,
        }
