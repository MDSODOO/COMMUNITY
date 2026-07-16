# -*- coding: utf-8 -*-
import base64
import io
import logging
import re
from datetime import date

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

_HEADER_DATE_RE = re.compile(
    r'INV\s+(\d{1,2})\s+DE\s+(\w+)\s+(\d{4})', re.IGNORECASE
)
_MONTH_ES_NUM = {
    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
}
_MONTH_ES_SHORT = {
    1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic',
}


class BiPosSupplierSnapshot(models.Model):
    _name = 'bi.pos.supplier.snapshot'
    _description = 'Snapshot de inventario de proveedor (BF)'
    _order = 'snapshot_date desc'

    name = fields.Char(string='Nombre', required=True)
    supplier_code = fields.Char(string='Clave proveedor', default='BF')
    snapshot_date = fields.Date(string='Fecha del inventario', required=True)
    active = fields.Boolean(default=True)
    file_attachment_id = fields.Many2one(
        'ir.attachment', string='Archivo Excel', ondelete='set null',
    )
    line_ids = fields.One2many(
        'bi.pos.supplier.snapshot.line', 'snapshot_id', string='Líneas',
    )
    line_count = fields.Integer(compute='_compute_line_count', string='# Productos')

    @api.depends('line_ids')
    def _compute_line_count(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)

    @api.model
    def import_from_xlsx(self, attachment_id):
        """Importa un Excel BF desde un ir.attachment ya subido.

        Detecta automáticamente la columna de inventario por su cabecera
        "INV DD DE MMM YYYY" y mapea las columnas de costo y precio por nombre.

        Returns:
            {'snapshot_id': int, 'name': str, 'imported': int, 'errors': int}
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('Instala openpyxl: pip install openpyxl'))

        attachment = self.env['ir.attachment'].browse(attachment_id)
        if not attachment.exists():
            raise UserError(_('Adjunto no encontrado.'))

        raw = base64.b64decode(attachment.datas)
        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)

        if 'INVENTARIO' not in wb.sheetnames:
            raise UserError(_('El Excel no contiene la hoja "INVENTARIO".'))

        ws = wb['INVENTARIO']
        rows = list(ws.iter_rows(min_row=7, values_only=True))
        if not rows:
            raise UserError(_('La hoja INVENTARIO está vacía.'))

        header_row = rows[0]

        # ── Detectar columna de inventario y su fecha ──────────────────
        inv_col = None
        inv_date = None
        for idx, cell in enumerate(header_row):
            m = _HEADER_DATE_RE.match(str(cell or ''))
            if m:
                month_abbr = m.group(2).upper()[:3]
                month_num = _MONTH_ES_NUM.get(month_abbr)
                if month_num:
                    inv_col = idx
                    inv_date = date(int(m.group(3)), month_num, int(m.group(1)))
                break

        if inv_col is None:
            raise UserError(_(
                'No se encontró la columna de inventario. '
                'El formato esperado es: "INV DD DE MMM YYYY" '
                '(ej. "INV 26 DE MAY 2026").'
            ))

        # ── Mapeo de columnas ──────────────────────────────────────────
        col_upper = {str(v or '').strip().upper(): i for i, v in enumerate(header_row)}

        def find_col(*candidates):
            for c in candidates:
                for k, i in col_upper.items():
                    if c in k:
                        return i
            return None

        codigo_col  = find_col('CODIGO')
        nombre_col  = find_col('NOMBRE COMERCIAL')
        barcode_col = find_col('CODIGO DE BARRAS')
        costo_col   = find_col('COSTO FACTURA CASA', 'COSTO FACTURA')
        precio_col  = find_col('PRECIO MOST', 'PRECIO MOSTRADOR')
        lab_col     = find_col('LABORATORIO')

        if codigo_col is None:
            raise UserError(_('No se encontró la columna "CODIGO" en la hoja INVENTARIO.'))

        # ── Crear cabecera del snapshot ────────────────────────────────
        label = f"{inv_date.day} {_MONTH_ES_SHORT[inv_date.month]} {inv_date.year}"
        snapshot = self.create({
            'name': f'BF {label}',
            'supplier_code': 'BF',
            'snapshot_date': inv_date,
            'file_attachment_id': attachment.id,
        })

        # ── Importar líneas ────────────────────────────────────────────
        def safe_float(row, col):
            if col is None:
                return 0.0
            try:
                return float(row[col] or 0)
            except (TypeError, ValueError):
                return 0.0

        def safe_str(row, col):
            if col is None:
                return ''
            return str(row[col] or '').strip()

        lines_vals = []
        errors = 0
        for row in rows[1:]:
            try:
                vendor_code = safe_str(row, codigo_col)
                if not vendor_code or vendor_code.upper() == 'NONE':
                    continue
                barcode = safe_str(row, barcode_col)
                lines_vals.append({
                    'snapshot_id': snapshot.id,
                    'vendor_code': vendor_code,
                    'barcode': barcode,
                    'product_name': safe_str(row, nombre_col),
                    'qty_on_hand': safe_float(row, inv_col),
                    'cost_price': safe_float(row, costo_col),
                    'retail_price': safe_float(row, precio_col),
                    'lab': safe_str(row, lab_col),
                })
            except Exception:
                errors += 1
                _logger.exception('BF snapshot: error procesando fila')
                continue

        self.env['bi.pos.supplier.snapshot.line'].create(lines_vals)

        _logger.info(
            'Snapshot BF "%s": %d líneas importadas, %d errores.',
            snapshot.name, len(lines_vals), errors,
        )
        return {
            'snapshot_id': snapshot.id,
            'name': snapshot.name,
            'imported': len(lines_vals),
            'errors': errors,
        }

    def get_pos_data(self):
        """Devuelve dict para join O(1) por barcode y por vendor_code."""
        self.ensure_one()
        result = {}
        for line in self.line_ids:
            entry = {
                'qty_on_hand': line.qty_on_hand,
                'cost_price': line.cost_price,
                'retail_price': line.retail_price,
            }
            if line.barcode:
                result[line.barcode] = entry
                stripped = line.barcode.lstrip('0')
                if stripped:
                    result[stripped] = entry
            if line.vendor_code:
                result[f'vc_{line.vendor_code}'] = entry
        return result

    def action_open_import_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'name': _('Importar Excel BF'),
            'res_model': 'import.bf.snapshot.wizard',
            'view_mode': 'form',
            'target': 'new',
        }


class BiPosSupplierSnapshotLine(models.Model):
    _name = 'bi.pos.supplier.snapshot.line'
    _description = 'Línea de snapshot de inventario BF'

    snapshot_id = fields.Many2one(
        'bi.pos.supplier.snapshot', ondelete='cascade', index=True, required=True,
    )
    vendor_code = fields.Char(string='Código BF', index=True)
    barcode = fields.Char(string='Código de barras', index=True)
    product_name = fields.Char(string='Nombre comercial')
    qty_on_hand = fields.Float(string='Inventario BF', digits=(12, 0))
    cost_price = fields.Float(string='Costo factura', digits=(12, 2))
    retail_price = fields.Float(string='Precio mostrador', digits=(12, 2))
    lab = fields.Char(string='Laboratorio')


class ImportBfSnapshotWizard(models.TransientModel):
    _name = 'import.bf.snapshot.wizard'
    _description = 'Asistente de importación de Excel BF'

    excel_file = fields.Binary(string='Archivo Excel (.xlsx)', required=True)
    filename = fields.Char(string='Nombre del archivo')

    def action_import(self):
        if not self.excel_file:
            raise UserError(_('Por favor selecciona un archivo Excel.'))

        attachment = self.env['ir.attachment'].create({
            'name': self.filename or 'inventario_bf.xlsx',
            'type': 'binary',
            'datas': self.excel_file,
            'res_model': 'bi.pos.supplier.snapshot',
        })

        result = self.env['bi.pos.supplier.snapshot'].import_from_xlsx(attachment.id)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Snapshot importado: %s', result['name']),
            'res_model': 'bi.pos.supplier.snapshot',
            'res_id': result['snapshot_id'],
            'view_mode': 'form',
            'target': 'current',
        }
